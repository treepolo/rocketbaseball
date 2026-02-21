import openpyxl

# Load with formulas
wb_formulas = openpyxl.load_workbook(r'd:\soft devloper\RSS\TrajectoryCalculator-new-3D-May2021.xlsx', data_only=False)
# Load with cached values
wb_values = openpyxl.load_workbook(r'd:\soft devloper\RSS\TrajectoryCalculator-new-3D-May2021.xlsx', data_only=True)

print("Sheet names:", wb_formulas.sheetnames)
print("="*80)

for sheet_name in wb_formulas.sheetnames:
    ws_f = wb_formulas[sheet_name]
    ws_v = wb_values[sheet_name]
    
    # Write each sheet to a separate file
    filename = f'd:\\soft devloper\\RSS\\sheet_{sheet_name.replace(" ","_").replace("/","_")}.txt'
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(f"SHEET: {sheet_name}\n")
        f.write(f"Dimensions: {ws_f.dimensions}\n")
        f.write(f"Max row: {ws_f.max_row}, Max col: {ws_f.max_column}\n")
        f.write("="*80 + "\n")
        
        for row in range(1, min(ws_f.max_row + 1, 800)):
            for col in range(1, min(ws_f.max_column + 1, 50)):
                cell_f = ws_f.cell(row=row, column=col)
                cell_v = ws_v.cell(row=row, column=col)
                
                if cell_f.value is not None:
                    coord = cell_f.coordinate
                    formula = cell_f.value
                    value = cell_v.value
                    
                    if isinstance(formula, str) and formula.startswith('='):
                        f.write(f"  {coord}: FORMULA={formula}  |  CACHED_VALUE={value}\n")
                    else:
                        f.write(f"  {coord}: {formula}\n")
    
    print(f"Written: {filename}")

# Named ranges
try:
    for name, defn in wb_formulas.defined_names.items():
        print(f"  Named range: {name} = {defn.attr_text}")
except Exception as e:
    print(f"Named ranges error: {e}")
    try:
        for defn in wb_formulas.defined_names.values():
            print(f"  Named range: {defn.name} = {defn.attr_text}")
    except Exception as e2:
        print(f"Named ranges error2: {e2}")

print("DONE")
